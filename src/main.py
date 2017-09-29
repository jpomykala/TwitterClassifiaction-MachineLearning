
import dataFacade
from textBlobProcessing import train_classifier
import os
import pprint
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter
import random
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from textUtils import clean_post
from datetime import date
from tweetsCorpra import get_corpus_ids
from twitterService import get_user_by_tweet_id
from twitterService import get_user_tweets
from twitterService import to_date

class Tweet:
	def __init__(self, tweet):
		self.text = tweet["text"]
		self.id = tweet["id"]
		self.date = tweet["created_at"]
		self.user_name = tweet["user"]["name"]
		self.user_id = tweet["user"]["id"]

excel_file_name = "data.xlsx"

row_id = 2
def iterate_over_tweets(tweets):
	global row_id 
	sick_start_date = None
	sick_post = None
	sick_post_id = None
	for tweet in tweets:
		text = clean_post(tweet.text)
		created = tweet.date
		classification = cl.classify(text)

		if classification is "sick" and sick_start_date is None:
			sick_start_date = to_date(created)
			sick_post = text
			continue

		if classification is "healthy" and sick_start_date is not None:
			cure_date = to_date(created)
			sick_duration_days = (cure_date - sick_start_date).days
			if sick_duration_days is 0:
				continue
			print("-------")
			print(tweet.user_name, "was sick at", sick_start, "because he/she wrote:", sick_post)
			print(tweet.user_name, "was cured at", cure_date, "because he/she wrote:", text)
			print(tweet.user_name, "was sick for", sick_duration_days, "days")
			print("-------")

			#save to xlsx
			excel_file = load_workbook(excel_file_name)
			sheet = excel_file.worksheets[0]
			sheet.cell(row=row_id, column=1).value = tweet.user_name # user name
			sheet.cell(row=row_id, column=2).value = tweet.user_id # twitter id
			sheet.cell(row=row_id, column=3).value = sick_start_date # sick start date
			sheet.cell(row=row_id, column=4).value = sick_post # sick post
			sheet.cell(row=row_id, column=5).value = sick_id # sick post id
			sheet.cell(row=row_id, column=6).value = cure_date # cured date
			sheet.cell(row=row_id, column=7).value = text # cured post 
			sheet.cell(row=row_id, column=8).value = tweet.id # cured post  id
			sheet.cell(row=row_id, column=9).value = sick_duration_days
			row_id = row_id + 1
			excel_file.save(excel_file_name)
			
			sick_start_date = None
			sick_post = None
			sick_post_id = None
			return

# dataFacade.update_data()

workbook = xlsxwriter.Workbook(excel_file_name)
worksheet = workbook.add_worksheet()
worksheet.write(0, 0, "user_name")
worksheet.write(0, 1, "user_id")
worksheet.write(0, 2, "sick_start")
worksheet.write(0, 3, "sick_post")
worksheet.write(0, 4, "sick_id")
worksheet.write(0, 5, "cure_date")
worksheet.write(0, 6, "cure_post")
worksheet.write(0, 7, "cure_id")
worksheet.write(0, 8, "duration")
workbook.close()

cl = train_classifier()
sick_tweet_ids = get_corpus_ids("sick")


print("Writing...")
for tweet_id in sick_tweet_ids:
	tweet_id = tweet_id.replace(".txt", "")
	try:
		user_id = get_user_by_tweet_id(int(tweet_id))
		tweets = get_user_tweets(int(user_id))
	except:
		continue
	else:
		tuple_tweets = [Tweet(t) for t in tweets]
		iterate_over_tweets(reversed(tuple_tweets))

