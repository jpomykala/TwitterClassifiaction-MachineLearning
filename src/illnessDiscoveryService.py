import os
import xlsxwriter
import random
import arrow
from pprint import pprint
from openpyxl import Workbook, load_workbook
from nltk import FreqDist
from textUtils import clean_post
from textblob import TextBlob
from tweetsCorpra import get_corpus_ids
from twitterService import get_user_by_tweet_id, get_user_tweets, to_date

excel_file_name = "results/data-" + arrow.now().format('YYYY-MM-DD HH:mm:ss') + ".xlsx"
def invoke(classifier, analyzer):
	create_output_file()
	print("--> Processing...")
	sick_tweet_ids = get_corpus_ids("sick")
	for tweet_id in sick_tweet_ids:
		tweet_id = tweet_id.replace(".txt", "")
		user_id = get_user_by_tweet_id(int(tweet_id))
		tweets = get_user_tweets(int(user_id))
		iterate_over_tweets(reversed(tweets), classifier, analyzer)
	return

def create_output_file():
	workbook = xlsxwriter.Workbook(excel_file_name)
	worksheet = workbook.add_worksheet()
	worksheet.write(0, 0, "user_name")
	worksheet.write(0, 1, "user_id")
	worksheet.write(0, 2, "sick_start")
	worksheet.write(0, 3, "sick_post")
	worksheet.write(0, 4, "sick_id")
	worksheet.write(0, 5, "cure_date")
	worksheet.write(0, 6, "cure_post")
	worksheet.write(0, 7, "cure_id")
	worksheet.write(0, 8, "duration")
	workbook.close()
	return

row_id = 2
def iterate_over_tweets(tweets, classifier, analyzer):
	global row_id 
	sick_start_date = None
	sick_post = None
	sick_post_id = None
	for tweet in tweets:
		text = TextBlob(clean_post(tweet.text), classifier=classifier, analyzer=analyzer)
		created = tweet.date

		if len(text.words) < 2:
			print("Less than 2 words")
			continue

		if text.detect_language() != "en":
			print("Not english posts")
			continue

		features = classifier.extract_features(text)
		classification = classifier.classify(text)

		print(classification, "|", text)
		prob_dist = classifier.prob_classify(text)
		print("prob:", prob_dist.max(), "sick:", prob_dist.prob("sick"),  "healthy:", prob_dist.prob("healthy"),  "neutral:", prob_dist.prob("neutral"))
		print("-------------")

		if classification == "sick" and sick_start_date is None:
			sick_start_date = to_date(created)
			sick_post = text
			sick_post_id = tweet.id
			continue

		if classification == "healthy" and sick_start_date is not None:
			cure_date = to_date(created)
			sick_duration_days = (cure_date - sick_start_date).days
			if sick_duration_days is 0:
				continue

			print("-------")
			print(tweet.user_name, "was sick at", sick_start_date, "because he/she wrote:", sick_post)
			print(tweet.user_name, "was cured at", cure_date, "because he/she wrote:", text)
			print(tweet.user_name, "was sick for", sick_duration_days, "days")
			print("-------")

			#save to xlsx
			excel_file = load_workbook(excel_file_name)
			sheet = excel_file.worksheets[0]
			sheet.cell(row=row_id, column=1).value = tweet.user_name # user name
			sheet.cell(row=row_id, column=2).value = tweet.user_id # twitter id
			sheet.cell(row=row_id, column=3).value = sick_start_date # sick start date
			sheet.cell(row=row_id, column=4).value = str(sick_post) # sick post
			sheet.cell(row=row_id, column=5).value = sick_post_id # sick post id
			sheet.cell(row=row_id, column=6).value = cure_date # cured date
			sheet.cell(row=row_id, column=7).value = str(text) # cured post 
			sheet.cell(row=row_id, column=8).value = tweet.id # cured post id
			sheet.cell(row=row_id, column=9).value = sick_duration_days
			row_id = row_id + 1
			excel_file.save(excel_file_name)
			
			sick_start_date = None
			sick_post = None
			sick_post_id = None
			return
